import pandas as pd
from openpyxl import load_workbook
from datetime import date, datetime
import os

# Separate Excel files for each data type
FILES = {
    "customer_details": "data/customer_details.xlsx",
    "store_details": "data/store_details.xlsx",
    "customer_bills": "data/customer_bills.xlsx",
    "distributor_details": "data/distributor_details.xlsx",
    "distributor_transactions": "data/distributor_transactions.xlsx",
    "credit_debit": "data/credit_debit.xlsx",
}

# Legacy support - redirect to separate files
FILE = FILES["customer_details"]  # Default for backward compatibility


# ---------------- HELPER FUNCTIONS ----------------

def _get_file(sheet_name):
    """Get the file path for a given sheet name"""
    return FILES.get(sheet_name, f"data/{sheet_name}.xlsx")


def _ensure_file_exists(file_path, columns):
    """Ensure an Excel file exists with the given columns, create if not"""
    if not os.path.exists(file_path):
        df = pd.DataFrame(columns=columns)
        df.to_excel(file_path, index=False)


def _normalize_header(h):
    if h is None:
        return ""
    return str(h).strip().lower().replace(" ", "_")


def normalize_columns(df):
    if df is None or df.empty:
        return df
    df.columns = (
        df.columns
        .astype(str)
        .str.strip()
        .str.lower()
        .str.replace(" ", "_")
    )
    return df


def _clean_value(val):
    """Clean pandas NaN/None values and return empty string"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, (int, float)) and pd.isna(val):
        return ""
    return val


# ---------------- READ FUNCTIONS ----------------

def read(sheet):
    """Read a sheet from its dedicated Excel file"""
    file_path = _get_file(sheet)
    try:
        return pd.read_excel(file_path, sheet_name=0)
    except:
        return pd.DataFrame()


def read_file(file_path, sheet_name=0):
    """Read from a specific file"""
    try:
        return pd.read_excel(file_path, sheet_name=sheet_name)
    except:
        return pd.DataFrame()


# ---------------- WRITE FUNCTIONS ----------------

def save_df(df, sheet_name):
    """Save a dataframe to its dedicated Excel file"""
    if df is None:
        return
    file_path = _get_file(sheet_name)
    df = df.fillna("")
    df.to_excel(file_path, index=False)


def _clean_row_for_excel(row):
    """Clean a row list to replace None/NaN with empty string for Excel"""
    cleaned = []
    for val in row:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            cleaned.append("")
        else:
            cleaned.append(val)
    return cleaned


def append(sheet, row):
    """Append a row to the sheet's Excel file using openpyxl"""
    file_path = _get_file(sheet)
    
    # Define column mappings for each sheet
    if sheet == "store_details":
        col_map = {
            "medicine": 1, "quantity": 2, "expiry_date": 3, "mfg_date": 4,
            "batch_number": 5, "mrp": 6, "distributor": 7, "purchase_date": 8
        }
    elif sheet == "customer_details":
        col_map = {"customer_id": 1, "customer_name": 2, "phone": 3, "address": 4}
    elif sheet == "distributor_details":
        col_map = {"distributor_id": 1, "distributor_name": 2, "phone_number": 3, "address": 4}
    elif sheet == "customer_bills":
        col_map = {"bill_no": 1, "customer_id": 2, "customer_name": 3, "date": 4, 
                   "total_amount": 5, "paid_amount": 6, "balance": 7}
    elif sheet == "distributor_transactions":
        col_map = {"receipt_no": 1, "bill_no": 2, "distributor_id": 3, "distributor_name": 4,
                   "phone_number": 5, "bill_amount": 6, "paid_amount": 7, "balance": 8, "date": 9}
    elif sheet == "credit_debit":
        col_map = {"customer_id": 1, "customer_name": 2, "payment_type": 3, "amount": 4, "date": 5, "note": 6}
    else:
        col_map = {}
    
    # Ensure file exists
    if not os.path.exists(file_path):
        if sheet == "customer_details":
            columns = ["customer_id", "customer_name", "phone", "address"]
        elif sheet == "store_details":
            columns = ["medicine", "quantity", "expiry_date", "mfg_date", "batch_number", "mrp", "distributor", "purchase_date"]
        elif sheet == "customer_bills":
            columns = ["bill_no", "customer_id", "customer_name", "date", "total_amount", "paid_amount", "balance"]
        elif sheet == "distributor_details":
            columns = ["distributor_id", "distributor_name", "phone_number", "address"]
        elif sheet == "distributor_transactions":
            columns = ["receipt_no", "bill_no", "distributor_id", "distributor_name", "phone_number", "bill_amount", "paid_amount", "balance", "date"]
        elif sheet == "credit_debit":
            columns = ["customer_id", "customer_name", "payment_type", "amount", "date", "note"]
        else:
            columns = []
        df = pd.DataFrame(columns=columns)
        df.to_excel(file_path, index=False)
    
    # Use openpyxl to append row in correct columns
    wb = load_workbook(file_path)
    ws = wb.active
    next_row = ws.max_row + 1
    
    for col_name, col_idx in col_map.items():
        if col_idx <= len(row):
            val = row[col_idx - 1]
            if val is None or (isinstance(val, float) and pd.isna(val)):
                val = ""
            ws.cell(next_row, col_idx, value=val)
    
    wb.save(file_path)


# ---------------- STORE DETAILS ----------------

def get_store_file():
    return FILES["store_details"]


def get_store_quantity(medicine, batch, expiry):
    """Get current quantity for a store entry"""
    file_path = FILES["store_details"]
    if not os.path.exists(file_path):
        return 0
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    headers = [_normalize_header(c.value) for c in ws[1]] if ws.max_row > 0 else []
    
    med_col = next((i+1 for i, h in enumerate(headers) if "medicine" in h), 1)
    batch_col = next((i+1 for i, h in enumerate(headers) if "batch" in h), 5)
    expiry_col = next((i+1 for i, h in enumerate(headers) if "expiry" in h), 3)
    qty_col = next((i+1 for i, h in enumerate(headers) if "quantity" in h or "qty" in h), 2)
    
    new_med = str(medicine).strip().lower()
    new_batch = str(batch).strip()
    new_expiry = str(expiry).strip()
    
    for r in range(2, ws.max_row + 1):
        med_val = ws.cell(r, med_col).value
        batch_val = ws.cell(r, batch_col).value
        expiry_val = ws.cell(r, expiry_col).value
        
        med_val_s = str(med_val).strip().lower() if med_val else ""
        batch_val_s = str(batch_val).strip() if batch_val else ""
        
        if isinstance(expiry_val, datetime):
            expiry_val_s = expiry_val.strftime("%Y-%m-%d")
        else:
            expiry_val_s = str(expiry_val).strip() if expiry_val else ""
        
        if med_val_s == new_med and batch_val_s == new_batch and expiry_val_s == new_expiry:
            qty = ws.cell(r, qty_col).value or 0
            try:
                return int(qty)
            except:
                return 0
    
    return 0


def save_store_entry(row):
    """Insert or update store entry."""
    file_path = FILES["store_details"]
    _ensure_file_exists(file_path, ["medicine", "quantity", "expiry_date", "mfg_date", "batch_number", "mrp", "distributor", "purchase_date"])
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Read headers
    headers = [_normalize_header(c.value) for c in ws[1]] if ws.max_row > 0 else []
    
    if not headers:
        headers = ["medicine", "quantity", "expiry_date", "mfg_date", "batch_number", "mrp", "distributor", "purchase_date"]
        ws.append(headers)
    
    # Find column indices
    med_col = next((i+1 for i, h in enumerate(headers) if "medicine" in h), 1)
    batch_col = next((i+1 for i, h in enumerate(headers) if "batch" in h), 5)
    expiry_col = next((i+1 for i, h in enumerate(headers) if "expiry" in h), 3)
    qty_col = next((i+1 for i, h in enumerate(headers) if "quantity" in h or "qty" in h), 2)
    
    new_med = str(row[0]).strip().lower()
    new_qty = int(row[1]) if row[1] else 0
    new_expiry = str(row[2]).strip()
    new_batch = str(row[4]).strip()
    
    # Prevent negative quantity on new entry
    if new_qty < 0:
        new_qty = 0
    
    # Search for existing entry
    for r in range(2, ws.max_row + 1):
        med_val = ws.cell(r, med_col).value
        batch_val = ws.cell(r, batch_col).value
        expiry_val = ws.cell(r, expiry_col).value
        
        med_val_s = str(med_val).strip().lower() if med_val else ""
        batch_val_s = str(batch_val).strip() if batch_val else ""
        expiry_val_s = str(expiry_val).strip() if expiry_val else ""
        
        if med_val_s == new_med and batch_val_s == new_batch and expiry_val_s == new_expiry:
            cur_qty = ws.cell(r, qty_col).value or 0
            try:
                cur_qty = int(cur_qty)
            except:
                cur_qty = 0
            
            ws.cell(r, qty_col, value=cur_qty + new_qty)
            wb.save(file_path)
            return "updated"
    
    # Append new entry
    ws.append(row)
    wb.save(file_path)
    return "inserted"


def find_store_entry(medicine, batch, expiry, distributor=None):
    """Find store entry"""
    file_path = FILES["store_details"]
    if not os.path.exists(file_path):
        return None
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    headers = [_normalize_header(c.value) for c in ws[1]] if ws.max_row > 0 else []
    
    med_col = next((i+1 for i, h in enumerate(headers) if "medicine" in h), 1)
    batch_col = next((i+1 for i, h in enumerate(headers) if "batch" in h), 5)
    expiry_col = next((i+1 for i, h in enumerate(headers) if "expiry" in h), 3)
    qty_col = next((i+1 for i, h in enumerate(headers) if "quantity" in h or "qty" in h), 2)
    
    new_med = str(medicine).strip().lower()
    new_batch = str(batch).strip()
    new_expiry = str(expiry).strip()
    
    for r in range(2, ws.max_row + 1):
        med_val = ws.cell(r, med_col).value
        batch_val = ws.cell(r, batch_col).value
        expiry_val = ws.cell(r, expiry_col).value
        
        med_val_s = str(med_val).strip().lower() if med_val else ""
        batch_val_s = str(batch_val).strip() if batch_val else ""
        expiry_val_s = str(expiry_val).strip() if expiry_val else ""
        
        if med_val_s == new_med and batch_val_s == new_batch and expiry_val_s == new_expiry:
            qty = ws.cell(r, qty_col).value or 0
            try:
                qty = int(qty)
            except:
                qty = 0
            return (r, qty)
    
    return None


def increment_store_quantity(medicine, batch, expiry, increment):
    """Increment store quantity - prevents negative values"""
    file_path = FILES["store_details"]
    if not os.path.exists(file_path):
        return False
    
    try:
        increment = int(increment)
    except:
        increment = 0
    
    # If decrementing, check if we have enough quantity
    if increment < 0:
        current_qty = get_store_quantity(medicine, batch, expiry)
        if current_qty + increment < 0:
            # Would result in negative quantity - set to 0 instead
            increment = -current_qty
            if increment == 0:
                return False  # No quantity to decrement
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    headers = [_normalize_header(c.value) for c in ws[1]] if ws.max_row > 0 else []
    
    med_col = next((i+1 for i, h in enumerate(headers) if "medicine" in h), 1)
    batch_col = next((i+1 for i, h in enumerate(headers) if "batch" in h), 5)
    expiry_col = next((i+1 for i, h in enumerate(headers) if "expiry" in h), 3)
    qty_col = next((i+1 for i, h in enumerate(headers) if "quantity" in h or "qty" in h), 2)
    
    new_med = str(medicine).strip().lower()
    new_batch = str(batch).strip()
    new_expiry = str(expiry).strip()
    
    for r in range(2, ws.max_row + 1):
        med_val = ws.cell(r, med_col).value
        batch_val = ws.cell(r, batch_col).value
        expiry_val = ws.cell(r, expiry_col).value
        
        med_val_s = str(med_val).strip().lower() if med_val else ""
        batch_val_s = str(batch_val).strip() if batch_val else ""
        
        if isinstance(expiry_val, datetime):
            expiry_val_s = expiry_val.strftime("%Y-%m-%d")
        else:
            expiry_val_s = str(expiry_val).strip() if expiry_val else ""
        
        if med_val_s == new_med and batch_val_s == new_batch and expiry_val_s == new_expiry:
            cur_qty = ws.cell(r, qty_col).value or 0
            try:
                cur_qty = int(cur_qty)
            except:
                cur_qty = 0
            
            new_qty = cur_qty + increment
            # Ensure quantity never goes negative
            if new_qty < 0:
                new_qty = 0
            
            ws.cell(r, qty_col, value=new_qty)
            wb.save(file_path)
            return True
    
    return False


def get_all_store_entries():
    file_path = FILES["store_details"]
    _ensure_file_exists(file_path, ["medicine", "quantity", "expiry_date", "manufacture_date", "batch_number", "mrp", "distributor", "purchase_date"])
    df = pd.read_excel(file_path)
    df = normalize_columns(df)
    
    if df.empty:
        return []
    
    records = []
    for rec in df.to_dict(orient="records"):
        def _get(*cands):
            for c in cands:
                if c in rec and rec[c] is not None:
                    val = rec[c]
                    if isinstance(val, float) and pd.isna(val):
                        return ""
                    return val
            return ""
        
        records.append({
            "medicine": str(_clean_value(_get("medicine", "medicine_name"))),
            "batch_number": str(_clean_value(_get("batch_number", "batch"))),
            "quantity": _get("quantity", "qty") or 0,
            "mrp": _clean_value(_get("mrp", "mrp_(₹)", "mrp_rs")),
            "distributor": str(_clean_value(_get("distributor", "distributor_name"))),
            "expiry_date": str(_clean_value(_get("expiry_date", "expiry"))),
            "mfg": str(_clean_value(_get("mfg", "manufacture_date", "manufacture"))),
            "purchase_date": str(_clean_value(_get("purchase_date", "dop", "date_of_purchase"))),
        })
    return records


def search_store_entries(query):
    """Search store entries"""
    file_path = FILES["store_details"]
    _ensure_file_exists(file_path, ["medicine", "quantity", "expiry_date", "mfg_date", "batch_number", "mrp", "distributor", "purchase_date"])
    df = pd.read_excel(file_path)
    df = normalize_columns(df)
    
    if df.empty:
        return []
    
    q = str(query).strip().lower()
    if not q:
        return get_all_store_entries()
    
    mask = pd.Series(False, index=df.index)
    for col in ["medicine", "batch_number", "distributor"]:
        if col in df.columns:
            mask = mask | df[col].astype(str).str.lower().str.contains(q, na=False)
    
    filtered = df[mask]
    
    records = []
    for rec in filtered.to_dict(orient="records"):
        def _get(*cands):
            for c in cands:
                if c in rec and rec[c] is not None:
                    val = rec[c]
                    if isinstance(val, float) and pd.isna(val):
                        return ""
                    return val
            return ""
        
        records.append({
            "medicine": str(_clean_value(_get("medicine", "medicine_name"))),
            "batch_number": str(_clean_value(_get("batch_number", "batch"))),
            "quantity": _get("quantity", "qty") or 0,
            "mrp": _clean_value(_get("mrp", "mrp_(₹)", "mrp_rs")),
            "distributor": str(_clean_value(_get("distributor", "distributor_name"))),
            "expiry_date": str(_clean_value(_get("expiry_date", "expiry"))),
            "mfg": str(_clean_value(_get("mfg", "mfg_date", "manufacture"))),
            "purchase_date": str(_clean_value(_get("purchase_date", "dop", "date_of_purchase"))),
        })
    return records


def get_expiring_products(months):
    """Get expiring products"""
    file_path = FILES["store_details"]
    _ensure_file_exists(file_path, ["medicine", "quantity", "expiry_date", "mfg_date", "batch_number", "mrp", "distributor", "purchase_date"])
    df = pd.read_excel(file_path)
    df = normalize_columns(df)
    
    if df.empty:
        return pd.DataFrame()
    
    df["expiry_date"] = pd.to_datetime(df["expiry_date"], errors="coerce")
    
    today = pd.to_datetime(date.today())
    future = today + pd.DateOffset(months=months)
    
    return df[
        (df["expiry_date"] >= today) &
        (df["expiry_date"] <= future)
    ][["medicine", "batch_number", "quantity", "expiry_date"]]


# ---------------- CUSTOMER DETAILS ----------------

def get_customer_file():
    return FILES["customer_details"]


def generate_customer_id():
    file_path = FILES["customer_details"]
    _ensure_file_exists(file_path, ["customer_id", "customer_name", "phone", "address"])
    df = pd.read_excel(file_path)
    df = normalize_columns(df)
    
    if df.empty:
        return "CUST001"
    
    last = int(str(df.iloc[-1]["customer_id"])[4:])
    return f"CUST{last + 1:03d}"


def add_customer(customer_id, name, phone, address):
    """Add a new customer"""
    file_path = FILES["customer_details"]
    _ensure_file_exists(file_path, ["customer_id", "customer_name", "phone", "address"])
    
    df = pd.read_excel(file_path)
    
    if not df.empty:
        existing = df[df["customer_id"] == customer_id]
        if not existing.empty:
            return False
    
    new_row = pd.DataFrame([{
        "customer_id": customer_id,
        "customer_name": str(name).strip().upper(),
        "phone": phone,
        "address": str(address).strip().upper()
    }])
    
    df = pd.concat([df, new_row], ignore_index=True)
    df = df.fillna("")
    df.to_excel(file_path, index=False)
    return True


def get_all_customers():
    file_path = FILES["customer_details"]
    _ensure_file_exists(file_path, ["customer_id", "customer_name", "phone", "address"])
    df = pd.read_excel(file_path)
    df = normalize_columns(df)
    
    if df.empty:
        return []
    
    records = []
    for rec in df.to_dict(orient="records"):
        records.append({
            "customer_id": rec.get("customer_id", ""),
            "customer_name": str(rec.get("customer_name", "")).upper(),
            "phone": rec.get("phone", ""),
            "address": str(rec.get("address", "")).upper(),
        })
    return records


def search_customers(query):
    """Search customers by name or ID"""
    file_path = FILES["customer_details"]
    _ensure_file_exists(file_path, ["customer_id", "customer_name", "phone", "address"])
    df = pd.read_excel(file_path)
    
    if df.empty:
        return []
    
    q = str(query).lower()
    filtered = df[
        df["customer_id"].astype(str).str.lower().str.contains(q) |
        df["customer_name"].astype(str).str.lower().str.contains(q)
    ]
    
    return [
        f"{row['customer_id']} / {row['customer_name']}"
        for _, row in filtered.iterrows()
    ]


def update_customer_by_id(customer_id, new_phone, new_address):
    """Update customer by ID"""
    file_path = FILES["customer_details"]
    _ensure_file_exists(file_path, ["customer_id", "customer_name", "phone", "address"])
    
    df = pd.read_excel(file_path)
    
    if df.empty:
        return False
    
    # Convert phone and address columns to string type to avoid dtype conflicts
    df["phone"] = df["phone"].astype(str)
    df["address"] = df["address"].astype(str)
    
    # Fill NaN values with empty strings
    df = df.fillna("")
    
    mask = df["customer_id"] == customer_id
    if not mask.any():
        return False
    
    df.loc[mask, "phone"] = str(new_phone).strip()
    df.loc[mask, "address"] = str(new_address).strip().upper()
    df.to_excel(file_path, index=False)
    return True


# ---------------- DISTRIBUTOR DETAILS ----------------

def get_distributor_file():
    return FILES["distributor_details"]


def generate_distributor_id():
    file_path = FILES["distributor_details"]
    _ensure_file_exists(file_path, ["distributor_id", "distributor_name", "phone_number", "address"])
    df = pd.read_excel(file_path)
    df = normalize_columns(df)
    
    if df.empty:
        return "DIST001"
    
    last = int(str(df.iloc[-1]["distributor_id"])[4:])
    return f"DIST{last + 1:03d}"


def distributor_name_exists(name):
    file_path = FILES["distributor_details"]
    _ensure_file_exists(file_path, ["distributor_id", "distributor_name", "phone_number", "address"])
    df = pd.read_excel(file_path)
    df = normalize_columns(df)
    
    if df.empty:
        return False
    
    return any(df["distributor_name"].astype(str).str.lower().eq(name.strip().lower()))


def add_distributor(dist_id, name, phone, address=""):
    """Add distributor"""
    file_path = FILES["distributor_details"]
    _ensure_file_exists(file_path, ["distributor_id", "distributor_name", "phone_number", "address"])
    
    if distributor_name_exists(name):
        return False
    
    df = pd.read_excel(file_path)
    
    new_row = pd.DataFrame([{
        "distributor_id": dist_id,
        "distributor_name": str(name).strip().upper(),
        "phone_number": phone,
        "address": str(address).strip().upper()
    }])
    
    df = pd.concat([df, new_row], ignore_index=True)
    df = df.fillna("")
    df.to_excel(file_path, index=False)
    return True


def get_all_distributors():
    file_path = FILES["distributor_details"]
    _ensure_file_exists(file_path, ["distributor_id", "distributor_name", "phone_number", "address"])
    df = pd.read_excel(file_path)
    df = normalize_columns(df)
    
    if df.empty:
        return []
    
    records = []
    for rec in df.to_dict(orient="records"):
        records.append({
            "distributor_id": rec.get("distributor_id", ""),
            "distributor_name": str(rec.get("distributor_name", "")).upper(),
            "phone_number": rec.get("phone_number", rec.get("phone", "")),
            "address": str(rec.get("address", rec.get("addr", ""))).upper(),
        })
    return records


def get_all_distributor_records():
    """Get all distributor records as list of dicts"""
    file_path = FILES["distributor_details"]
    _ensure_file_exists(file_path, ["distributor_id", "distributor_name", "phone_number", "address"])
    df = pd.read_excel(file_path)
    df = normalize_columns(df)
    
    if df.empty:
        return []
    
    records = []
    for rec in df.to_dict(orient="records"):
        records.append({
            "distributor_id": rec.get("distributor_id", ""),
            "distributor_name": str(rec.get("distributor_name", "")).upper(),
            "phone_number": rec.get("phone_number", rec.get("phone", "")),
            "address": str(rec.get("address", rec.get("addr", ""))).upper(),
        })
    return records


def get_distributor_by_name(name):
    file_path = FILES["distributor_details"]
    _ensure_file_exists(file_path, ["distributor_id", "distributor_name", "phone_number", "address"])
    df = pd.read_excel(file_path)
    df = normalize_columns(df)
    
    row = df[df["distributor_name"] == name]
    
    if row.empty:
        return None
    
    r = row.iloc[0]
    return [r["distributor_id"], r["distributor_name"], r["phone_number"]]


def update_distributor_phone(name, phone, new_phone, new_address=None):
    """Update distributor phone"""
    file_path = FILES["distributor_details"]
    _ensure_file_exists(file_path, ["distributor_id", "distributor_name", "phone_number", "address"])
    
    df = pd.read_excel(file_path)
    
    if df.empty:
        return 0
    
    name_lower = str(name).lower()
    mask = df["distributor_name"].astype(str).str.lower().str.contains(name_lower, na=False)
    
    if not mask.any():
        return 0
    
    if new_phone:
        df.loc[mask, "phone_number"] = new_phone
    if new_address:
        df.loc[mask, "address"] = str(new_address).strip().upper()
    
    df = df.fillna("")
    df.to_excel(file_path, index=False)
    return mask.sum()


def update_distributor_by_id(distributor_id, new_phone_number, new_address):
    """Update distributor by ID"""
    file_path = FILES["distributor_details"]
    _ensure_file_exists(file_path, ["distributor_id", "distributor_name", "phone_number", "address"])
    
    df = pd.read_excel(file_path)
    
    if df.empty:
        return False
    
    # Convert phone and address columns to string type to avoid dtype conflicts
    df["phone_number"] = df["phone_number"].astype(str)
    df["address"] = df["address"].astype(str)
    
    # Fill NaN values with empty strings
    df = df.fillna("")
    
    df["distributor_id"] = df["distributor_id"].astype(str)
    distributor_id = str(distributor_id)
    
    mask = df["distributor_id"] == distributor_id
    if not mask.any():
        return False
    
    if new_phone_number:
        df.loc[mask, "phone_number"] = str(new_phone_number).strip()
    if new_address:
        df.loc[mask, "address"] = str(new_address).strip().upper()
    
    df.to_excel(file_path, index=False)
    return True


def search_distributors(query):
    """Search distributors"""
    file_path = FILES["distributor_details"]
    _ensure_file_exists(file_path, ["distributor_id", "distributor_name", "phone_number", "address"])
    df = pd.read_excel(file_path)
    df = normalize_columns(df)
    
    if df.empty:
        return []
    
    q = str(query).strip().lower()
    if not q:
        return df["distributor_name"].unique().tolist()
    
    filtered = df[df["distributor_name"].astype(str).str.lower().str.contains(q, na=False)]
    return filtered["distributor_name"].unique().tolist()


def delete_distributor_and_transactions(distributor_id=None, distributor_name=None):
    """Delete distributor and their transactions"""
    deleted_details = 0
    deleted_tx = 0
    
    file_path = FILES["distributor_details"]
    if os.path.exists(file_path):
        df = pd.read_excel(file_path)
        
        if distributor_id:
            initial_len = len(df)
            df = df[df["distributor_id"] != distributor_id]
            deleted_details = initial_len - len(df)
        elif distributor_name:
            name_lower = str(distributor_name).lower()
            initial_len = len(df)
            df = df[~df["distributor_name"].astype(str).str.lower().str.contains(name_lower, na=False)]
            deleted_details = initial_len - len(df)
        
        df = df.fillna("")
        df.to_excel(file_path, index=False)
    
    tx_file = FILES["distributor_transactions"]
    if os.path.exists(tx_file):
        df = pd.read_excel(tx_file)
        
        if distributor_id:
            initial_len = len(df)
            df = df[df["distributor_id"] != distributor_id]
            deleted_tx = initial_len - len(df)
        elif distributor_name:
            name_lower = str(distributor_name).lower()
            initial_len = len(df)
            df = df[~df["distributor_name"].astype(str).str.lower().str.contains(name_lower, na=False)]
            deleted_tx = initial_len - len(df)
        
        df = df.fillna("")
        df.to_excel(tx_file, index=False)
    
    return (deleted_details, deleted_tx)


# ---------------- DISTRIBUTOR TRANSACTIONS ----------------

def get_distributor_tx_file():
    return FILES["distributor_transactions"]


def receipt_exists(receipt_no):
    """Check if receipt exists"""
    file_path = FILES["distributor_transactions"]
    if not os.path.exists(file_path):
        return False
    
    df = pd.read_excel(file_path)
    if df.empty:
        return False
    
    return any(df["receipt_no"].astype(str).str.strip() == str(receipt_no).strip())


def get_distributor_transactions(distributor_name=None):
    """Get distributor transactions"""
    file_path = FILES["distributor_transactions"]
    _ensure_file_exists(file_path, ["receipt_no", "bill_no", "distributor_id", "distributor_name", "phone_number", "bill_amount", "paid_amount", "balance", "date"])
    
    df = pd.read_excel(file_path)
    df = normalize_columns(df)
    
    if df.empty:
        return []
    
    rows = []
    for rec in df.to_dict(orient="records"):
        bill_amt = rec.get("bill_amount", 0) or 0
        paid_amt = rec.get("paid_amount", 0) or 0
        bal = rec.get("balance")
        
        try:
            bill_n = float(bill_amt) if bill_amt else 0.0
        except:
            bill_n = 0.0
        
        try:
            paid_n = float(paid_amt) if paid_amt else 0.0
        except:
            paid_n = 0.0
        
        if bal is None or bal == "":
            bal_n = bill_n - paid_n
        else:
            try:
                bal_n = float(bal)
            except:
                bal_n = bill_n - paid_n
        
        row = {
            "receipt_no": rec.get("receipt_no", ""),
            "bill_no": rec.get("bill_no", ""),
            "distributor_id": rec.get("distributor_id", ""),
            "distributor_name": str(rec.get("distributor_name", "")),
            "phone_number": rec.get("phone_number", ""),
            "bill_amount": bill_n,
            "paid_amount": paid_n,
            "balance": bal_n,
            "date": rec.get("date", ""),
        }
        rows.append(row)
    
    if distributor_name:
        q = str(distributor_name).strip().lower()
        rows = [r for r in rows if q in str(r.get("distributor_name", "")).lower()]
    
    return rows


def get_distributor_pending_sums():
    """Get distributor pending sums"""
    file_path = FILES["distributor_transactions"]
    if not os.path.exists(file_path):
        return []
    
    df = pd.read_excel(file_path)
    df = normalize_columns(df)
    
    if df.empty:
        return []
    
    name_col = "distributor_name" if "distributor_name" in df.columns else "distributor"
    bal_col = "balance" if "balance" in df.columns else None
    
    if bal_col is None:
        if "bill_amount" in df.columns and "paid_amount" in df.columns:
            df["balance_calc"] = pd.to_numeric(df["bill_amount"], errors="coerce").fillna(0) - pd.to_numeric(df["paid_amount"], errors="coerce").fillna(0)
            bal_col = "balance_calc"
        else:
            return []
    
    if name_col not in df.columns:
        return []
    
    grp = df.groupby(df[name_col].astype(str).str.strip()).agg({bal_col: lambda s: pd.to_numeric(s, errors="coerce").fillna(0).sum()})
    
    return [
        {"distributor_name": name, "total_pending": float(row[bal_col])}
        for name, row in grp.iterrows()
    ]


def get_distributor_pending_summary(search_text=None):
    """Get distributor pending summary"""
    file_path = FILES["distributor_transactions"]
    if not os.path.exists(file_path):
        return []
    
    df = pd.read_excel(file_path)
    df = normalize_columns(df)
    
    if df.empty:
        return []
    
    name_col = "distributor_name" if "distributor_name" in df.columns else "distributor"
    if name_col not in df.columns:
        return []
    
    bill = pd.to_numeric(df.get("bill_amount", pd.Series(0)), errors="coerce").fillna(0)
    paid = pd.to_numeric(df.get("paid_amount", pd.Series(0)), errors="coerce").fillna(0)
    df["balance_val"] = bill - paid
    df["dist_norm"] = df[name_col].astype(str).str.strip().str.upper()
    
    if search_text:
        q = str(search_text).strip().upper()
        df = df[df["dist_norm"].str.contains(q, na=False)]
    
    grp = df.groupby("dist_norm", as_index=False)["balance_val"].sum()
    
    return [
        {"distributor_name": row["dist_norm"], "total_pending": float(row["balance_val"])}
        for _, row in grp.iterrows()
    ]


def get_distributor_names_from_transactions(search_text=None):
    """Get distributor names from transactions"""
    file_path = FILES["distributor_transactions"]
    if not os.path.exists(file_path):
        return []
    
    df = pd.read_excel(file_path)
    df = normalize_columns(df)
    
    if df.empty:
        return []
    
    col = "distributor_name" if "distributor_name" in df.columns else "distributor"
    if col not in df.columns:
        return []
    
    names = df[col].astype(str).str.strip().str.upper().dropna().unique().tolist()
    
    if search_text:
        q = str(search_text).strip().upper()
        names = [n for n in names if q in n]
    
    return sorted(names)


def add_distributor_adjustment(distributor_id, distributor_name, kind, amount, note=None):
    """Add distributor adjustment"""
    file_path = FILES["distributor_transactions"]
    _ensure_file_exists(file_path, ["receipt_no", "bill_no", "distributor_id", "distributor_name", "phone_number", "bill_amount", "paid_amount", "balance", "date"])
    
    df = pd.read_excel(file_path)
    
    new_row = pd.DataFrame([{
        "receipt_no": "",
        "bill_no": "",
        "distributor_id": distributor_id,
        "distributor_name": distributor_name,
        "phone_number": "",
        "bill_amount": 0 if kind == "debit" else amount,
        "paid_amount": amount if kind == "debit" else 0,
        "balance": -amount if kind == "debit" else amount,
        "date": date.today()
    }])
    
    df = pd.concat([df, new_row], ignore_index=True)
    df = df.fillna("")
    df.to_excel(file_path, index=False)
    return True


# ---------------- CREDIT DEBIT ----------------

def get_credit_debit_file():
    return FILES["credit_debit"]


def get_customer_summary(customer_id, customer_name):
    """Get customer credit/debit summary"""
    file_path = FILES["credit_debit"]
    _ensure_file_exists(file_path, ["customer_id", "customer_name", "payment_type", "amount", "date", "note"])
    
    try:
        df = pd.read_excel(file_path)
        df.columns = df.columns.str.strip()
        
        customer_df = df[
            (df["customer_id"].astype(str) == str(customer_id)) &
            (df["customer_name"].astype(str).str.lower() == str(customer_name).lower())
        ]
        
        if customer_df.empty:
            return 0, 0, 0
        
        payment_column = "payment_type"
        
        total_credit = customer_df.loc[
            customer_df[payment_column].str.upper() == "CREDIT", "amount"
        ].sum()
        
        total_debit = customer_df.loc[
            customer_df[payment_column].str.upper() == "DEBIT", "amount"
        ].sum()
        
        balance = total_credit - total_debit
        
        return float(total_credit), float(total_debit), float(balance)
    except Exception as e:
        print("Error in get_customer_summary:", e)
        return 0, 0, 0


def search_customers_credit(keyword):
    """Search customers in credit_debit"""
    file_path = FILES["credit_debit"]
    if not os.path.exists(file_path):
        return []
    
    df = pd.read_excel(file_path)
    if df.empty:
        return []
    
    results = set()
    for _, row in df.iterrows():
        cname = str(row.get("customer_name", ""))
        if keyword.lower() in cname.lower():
            results.add((row.get("customer_id", ""), cname))
    
    return list(results)


# ---------------- BILLING ----------------

def get_customer_bills_file():
    return FILES["customer_bills"]


def add_bill_record(bill_no, customer_id, customer_name, date, total_amount, paid_amount, balance):
    """Add a bill record"""
    file_path = FILES["customer_bills"]
    _ensure_file_exists(file_path, ["bill_no", "customer_id", "customer_name", "date", "total_amount", "paid_amount", "balance"])
    
    df = pd.read_excel(file_path)
    
    new_row = pd.DataFrame([{
        "bill_no": bill_no,
        "customer_id": customer_id,
        "customer_name": customer_name,
        "date": date,
        "total_amount": total_amount,
        "paid_amount": paid_amount,
        "balance": balance
    }])
    
    df = pd.concat([df, new_row], ignore_index=True)
    df = df.fillna("")
    df.to_excel(file_path, index=False)


def get_customer_bills(customer_id=None, customer_name=None):
    """Get customer bills"""
    file_path = FILES["customer_bills"]
    _ensure_file_exists(file_path, ["bill_no", "customer_id", "customer_name", "date", "total_amount", "paid_amount", "balance"])
    
    df = pd.read_excel(file_path)
    
    if customer_id:
        df = df[df["customer_id"].astype(str) == str(customer_id)]
    if customer_name:
        df = df[df["customer_name"].astype(str).str.lower().str.contains(str(customer_name).lower(), na=False)]
    
    return df.to_dict(orient="records")


# ---------------- ADDITIONAL HELPER FUNCTIONS ----------------

def update_customer_address(name, phone, new_address):
    """Update customer address"""
    file_path = FILES["customer_details"]
    _ensure_file_exists(file_path, ["customer_id", "customer_name", "phone", "address"])
    
    df = pd.read_excel(file_path)
    
    if df.empty:
        return 0
    
    # Convert phone and address columns to string type to avoid dtype conflicts
    df["phone"] = df["phone"].astype(str)
    df["address"] = df["address"].astype(str)
    
    # Fill NaN values with empty strings
    df = df.fillna("")
    
    name_lower = str(name).lower()
    mask = df["customer_name"].astype(str).str.lower().str.contains(name_lower, na=False)
    
    if phone:
        phone_str = str(phone).strip()
        mask = mask & df["phone"].astype(str).str.contains(phone_str, na=False)
    
    if not mask.any():
        return 0
    
    df.loc[mask, "address"] = str(new_address).strip().upper()
    df.to_excel(file_path, index=False)
    return mask.sum()
