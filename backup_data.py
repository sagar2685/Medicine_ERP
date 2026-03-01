"""
================================================================================
              DATA BACKUP AND CLEAR SCRIPT
================================================================================

This script backs up all data files and clears the data while keeping headers.
Run: python backup_data.py

It will:
1. Create a backup folder: data_YYYYMMDD_backup
2. Copy all Excel files to the backup folder
3. Clear all data rows from the original files (keeping headers only)

The backup folder can be used to restore data if needed.
================================================================================
"""

import os
import shutil
from datetime import datetime

# Try to import pandas, if not available use openpyxl
try:
    import pandas as pd
    USE_PANDAS = True
except ImportError:
    from openpyxl import load_workbook
    USE_PANDAS = False


def backup_and_clear_data():
    """Backup data files and clear all data except headers"""
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(script_dir, "data")
    
    if not os.path.exists(data_dir):
        print("ERROR: Data folder not found!")
        return False
    
    # Create backup folder with current date
    current_date = datetime.now().strftime("%Y%m%d")
    backup_folder_name = f"data_{current_date}_backup"
    backup_dir = os.path.join(script_dir, backup_folder_name)
    
    # Check if backup folder already exists
    if os.path.exists(backup_dir):
        print(f"WARNING: Backup folder '{backup_folder_name}' already exists.")
        response = input("Do you want to overwrite it? (y/n): ")
        if response.lower() != 'y':
            print("Cancelled.")
            return False
        shutil.rmtree(backup_dir)
    
    # Create backup directory
    os.makedirs(backup_dir)
    print(f"Created backup folder: {backup_folder_name}/")
    
    # Get list of Excel files in data folder
    excel_files = [f for f in os.listdir(data_dir) if f.endswith('.xlsx')]
    
    if not excel_files:
        print("ERROR: No Excel files found in data folder!")
        return False
    
    print(f"\nProcessing {len(excel_files)} file(s)...\n")
    
    for filename in excel_files:
        source_file = os.path.join(data_dir, filename)
        backup_file = os.path.join(backup_dir, filename)
        
        try:
            if USE_PANDAS:
                # Using pandas
                df = pd.read_excel(source_file)
                
                if not df.empty:
                    # Save to backup
                    df.to_excel(backup_file, index=False)
                    print(f"  Backed up: {filename} ({len(df)} rows)")
                    
                    # Clear data but keep headers
                    df_empty = df.head(0)  # Get only headers
                    df_empty.to_excel(source_file, index=False)
                    print(f"  Cleared: {filename}")
                else:
                    # Empty file - just backup
                    df.to_excel(backup_file, index=False)
                    print(f"  Backed up (empty): {filename}")
            else:
                # Using openpyxl
                wb = load_workbook(source_file)
                ws = wb.active
                
                # Count rows (excluding header)
                row_count = ws.max_row - 1 if ws.max_row > 1 else 0
                
                if row_count > 0:
                    # Save current data to backup
                    wb.save(backup_file)
                    print(f"  Backed up: {filename} ({row_count} rows)")
                    
                    # Clear all data rows (keep row 1 which is header)
                    ws.delete_rows(2, ws.max_row)
                    wb.save(source_file)
                    print(f"  Cleared: {filename}")
                else:
                    # Empty file - just backup
                    wb.save(backup_file)
                    print(f"  Backed up (empty): {filename}")
                    
        except Exception as e:
            print(f"  ERROR processing {filename}: {e}")
    
    print(f"\n{'='*60}")
    print(f"SUCCESS!")
    print(f"  Backup location: {backup_folder_name}/")
    print(f"  Files processed: {len(excel_files)}")
    print(f"{'='*60}")
    
    return True


if __name__ == "__main__":
    print("=" * 60)
    print("       DATA BACKUP AND CLEAR SCRIPT")
    print("=" * 60)
    print()
    print("This will:")
    print("  1. Create a backup folder with all current data")
    print("  2. Clear all data from Excel files (keep headers)")
    print()
    print("-" * 60)
    
    response = input("Continue? (y/n): ")
    if response.lower() == 'y':
        print()
        backup_and_clear_data()
    else:
        print("\nCancelled.")
