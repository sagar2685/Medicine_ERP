from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from datetime import datetime
import os
import re
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "data/erp.xlsx"  # Excel file to store bills
SHEET_NAME = "customer_bills"

def save_bill_record(bill_id, customer_name, date, total_amount, pdf_name):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        if SHEET_NAME in wb.sheetnames:
            sheet = wb[SHEET_NAME]
        else:
            # create sheet and add header
            sheet = wb.create_sheet(SHEET_NAME)
            sheet.append(["bill_id", "customer_name", "date", "total_amount", "pdf_name"])
    else:
        # create new workbook and sheet
        wb = Workbook()
        sheet = wb.active
        sheet.title = SHEET_NAME
        sheet.append(["bill_id", "customer_name", "date", "total_amount", "pdf_name"])

    # Append bill data
    sheet.append([bill_id, customer_name, date, total_amount, pdf_name])
    wb.save(EXCEL_FILE)
    wb.close()  # Ensure the file is properly closed



def extract_customer_name(x):
    """Extract customer name from format 'CUST001 / John Smith'"""
    if " / " in x:
        return x.split(" / ", 1)[1].strip()
    return x.strip()

def generate_bill(customer_name, items):
    """
    Generate PDF bill for multiple items and save record to Excel.
    items: list of tuples -> (medicine_name, expiry, batch, qty, mrp, total)
    Bills are saved in date-based subfolders (e.g., customer_bills/2026-03-01/)
    """
    customer_name_only = extract_customer_name(customer_name)
    # Create safe filename - remove invalid characters and replace spaces with hyphens
    safe_name = customer_name_only.replace(" ", "-")
    safe_name = re.sub(r'[<>:"\\|?*]', '', safe_name)
    # Remove any leading/trailing hyphens
    safe_name = safe_name.strip('-')
    
    # Create date-based subfolder
    today_date = datetime.now().strftime('%Y-%m-%d')
    bills_folder = os.path.join("customer_bills", today_date)
    if not os.path.exists(bills_folder):
        os.makedirs(bills_folder)

    bill_no = f"BILL-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    filename = f"{safe_name}_{bill_no}.pdf"
    path = os.path.join(bills_folder, filename)

    c = canvas.Canvas(path, pagesize=A4)
    width, height = A4

    # ---------------- HEADER ----------------
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(width / 2, height - 40, "GITA MEDICAL HALL")
    c.line(30, height - 55, width - 30, height - 55)

    # STORE INFO
    c.setFont("Helvetica", 10)
    c.drawString(40, height - 80, "Lowairpoa, Dist. Karimganj, Assam - 788726")
    c.drawString(40, height - 95, "Phone: 9854709080")

    # BILL INFO
    c.drawRightString(width - 40, height - 80, f"Bill No: {bill_no}")
    c.drawRightString(width - 40, height - 95, f"Date: {datetime.now().strftime('%d-%m-%Y')}")

    # CUSTOMER
    
    c.setFont("Helvetica-Bold", 11)
    c.drawString(40, height - 130, "Customer:")
    c.setFont("Helvetica", 10)
    c.drawString(120, height - 130, customer_name_only)

    # ---------------- TABLE HEADER ----------------
    c.setFont("Helvetica-Bold", 10)
    headers = ["Medicine", "Expiry", "Qty", "MRP", "Total"]
    x_positions = [40, 180, 260, 340, 420]

    for header, x in zip(headers, x_positions):
        c.drawString(x, height - 150, header)

    # ---------------- TABLE CONTENT ----------------
    c.setFont("Helvetica", 10)
    y = height - 170
    grand_total = 0

    for item in items:
        medicine, expiry, batch, qty, mrp, total = item
        values = [medicine, expiry, str(qty), f"{mrp:.2f}", f"{total:.2f}"]
        for val, x in zip(values, x_positions):
            c.drawString(x, y, val)
        y -= 20
        grand_total += total

        if y < 100:  # Add new page if overflow
            c.showPage()
            y = height - 50

    # ---------------- GRAND TOTAL ----------------
    c.setFont("Helvetica-Bold", 12)
    c.drawRightString(width - 40, y - 20, f"Grand Total: ₹ {grand_total:.2f}")

    # ---------------- FOOTER ----------------
    c.setFont("Helvetica", 9)
    c.drawCentredString(width / 2, 50, "Thank you for your purchase!")

    c.save()

    # ---------------- SAVE RECORD ----------------
    save_bill_record(
        bill_id=bill_no,
        customer_name=customer_name,
        date = datetime.now().strftime('%d-%m-%Y %H:%M:%S'),
        total_amount=grand_total,
        pdf_name=filename
    )

    return filename
